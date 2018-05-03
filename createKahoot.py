#!/usr/bin/env python2

import sys
import time
import re
import os
import openpyxl
import selenium
import gtk, pygtk

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains

driver = webdriver.Chrome("/usr/lib/chromium-browser/chromedriver")
# Open camera ip page
driver.get("https://kahoot.com/welcomeback/")
actions=ActionChains(driver)

#relocate window
driver.set_window_size(gtk.gdk.screen_width()/2, gtk.gdk.screen_height())
driver.set_window_position(0, 0)

def login():
	time.sleep(1)
	button=driver.find_element_by_xpath('//*[@id="menu-header-menu"]/li[4]/a')
	button.click()
	time.sleep(1)
	username=driver.find_element_by_xpath('//*[@id="username-input-field"]/div[1]/input')
	username.send_keys("TeachBot")
	password=driver.find_element_by_xpath('//*[@id="password-input-field"]/div[1]/input')
	password.send_keys("IamTeachBot")
	signin=driver.find_element_by_xpath('//*[@id="layout"]/div/main/div/form/button[2]')
	signin.click()
	return

def createKahoot():
	button=driver.find_element_by_xpath('//*[@id="top-menu-bar"]/div[2]/div/a')
	button.click()
	button=driver.find_element_by_xpath('//*[@id="screen"]/div/div/div/section/div/ul/li[1]/a')
	button.click()
	return

def testSummary(filename):
	title=driver.find_element_by_xpath('//*[@id="layout"]/div/main/form/div/div/div[1]/div/div/input')
	title.send_keys(filename)
	summary=driver.find_element_by_xpath('//*[@id="layout"]/div/main/form/div/div/div[2]/div/div/textarea')
	summary.send_keys('Automated test')

	button=driver.find_element_by_xpath('//*[@id="layout"]/div/main/form/div[2]/div[3]/div/div/div/select')
	button.click()
	button=driver.find_element_by_xpath('//*[@id="layout"]/div/main/form/div[2]/div[3]/div/div/div/select/option[2]')
	button.click()

	button=driver.find_element_by_xpath('//*[@id="app"]/div/div/span/section/div/div/div/div/button[1]')
	button.click()
	time.sleep(1)
	button=driver.find_element_by_xpath('//*[@id="layout"]/div/header/header/div[3]/button[1]')
	button.click()
	button=driver.find_element_by_xpath('//*[@id="app"]/div/div/span/section/div/div/div/div/button[1]')
	button.click()
	time.sleep(2)
	button=driver.find_element_by_xpath('//*[@id="layout"]/div/header/header/div[3]/button[1]')
	button.click()
	return

def addQuestion(ws, i):
	actions= ActionChains(driver)

	button=driver.find_element_by_xpath('//*[@id="layout"]/div/main/section[2]/div/div/button')
	button.click()
	time.sleep(1)

	#Send question
	question=driver.find_element_by_xpath('//*[@id="ql-editor-1"]/div')
	question.click()
	actions.send_keys(str(ws.cell(row=i , column=2 ).value))
	actions.perform() 
	actions.reset_actions()
	j=3
	x=1
	while ws.cell(row=i , column=j).value!=None:
		#Send answers
		answer=driver.find_element_by_xpath('//*[@id="ql-editor-'+str(x+1)+'"]/div')
		answer.click()
		actions.send_keys(str(ws.cell(row=i , column=j ).value))
		actions.perform()
		actions.reset_actions()
		#Validate answer
		if str(ws.cell(row=i , column=j+1 ).value)=='correct' or str(ws.cell(row=i , column=j+1 ).value)=='correcto':
			if (x)%2==0:
				b=2
			else:
				b=1
			if x<3:
				a=1
			else:	
				a=2
			answer2=driver.find_element_by_xpath('//*[@id="layout"]/div/main/form/div[2]/div['+str(a)+']/div['+str(b)+']/div/div/div[2]/label/span')
			answer2.click()
		j=j+2
		x=x+1

	button=driver.find_element_by_xpath('//*[@id="layout"]/div/header/header/div[3]/button')
	button.click()
	return

def obtain_key ():
	text=driver.find_element_by_xpath('//*[@id="mainView"]/div[2]/div[2]/div[1]/div[1]/div[2]/span/strong')	

def check_exists_by_xpath(xpath):
	try:
		driver.find_element_by_xpath(xpath)
	except NoSuchElementException:
		return False
	return True
	
def main():
	filename="Test"
	#logn to server
	print("loging to kahoot...")
	time.sleep(2)		
	login()
	time.sleep(2)
	print("creating new kahoot...")
	createKahoot()
	time.sleep(2)
	
	#Questions
	print("adding questions...")
	wb = openpyxl.load_workbook(filename+'.xlsx')
	ws = wb.active

	testSummary(filename)
	for i in range (1, ws.max_row+1):
		time.sleep(1)
		addQuestion(ws, i)
	#Start quiz
	print("starting kahoot...")
	button=driver.find_element_by_xpath('//*[@id="layout"]/div/header/header/div[3]/button')
	button.click()
	time.sleep(3)
	button=driver.find_element_by_xpath('//*[@id="layout"]/div/header/header/div[3]/button')
	button.click()
	time.sleep(3)
	button=driver.find_element_by_xpath('//*[@id="kahoot-list-container"]/div/table[2]/tbody/tr/td[4]/button')
	button.click()
	time.sleep(15)
	driver.close()
	tabs=driver.window_handles
	driver.switch_to.window(str(tabs[0]))
	
	button=driver.find_element_by_xpath('//*[@id="mainView"]/div[3]/div/div[1]/div[1]/button')
	button.click()
	time.sleep(7)

	#Obtain key and start players
	text=driver.find_element_by_xpath('//*[@id="mainView"]/div[2]/div[2]/div[1]/div[1]/div[2]/span/strong')	
	print (text.text)

	print("starting players...")
	for i in range (0,2):
		run='python player1.py '+text.text+' '+str(i)+' '+str(ws.max_row)
		os.system("gnome-terminal -e "+"'"+run+"'")

	print("Lets Go!!")
	#Wait for players and start quiz
	time.sleep(15)
	button=driver.find_element_by_xpath('//*[@id="mainView"]/div[3]/button')
	button.click()

	time.sleep(10)
	
	#Wait for question to finish
	for i in range (0, ws.max_row):
		time.sleep(30)
		iframe=driver.find_element_by_css_selector('iframe')
		driver.switch_to_frame(iframe)
		button=driver.find_element_by_xpath('//*[@id="app"]/main/div[1]/section[1]/section[3]/div/aside/button')
		button.click()

		#return to original frame
		driver.switch_to_default_content()

		if i!=ws.max_row-1:
			time.sleep(3)
			iframe=driver.find_element_by_id('scoreBoardIframe')
			driver.switch_to_frame(iframe)
			button=driver.find_element_by_xpath('//*[@id="app"]/main/div[1]/div/div/button')
			button.click()

			driver.switch_to_default_content()

	#Get results
	time.sleep(5)
	iframe=driver.find_element_by_id('scoreBoardIframe')
	driver.switch_to_frame(iframe)
	#button=driver.find_element_by_xpath('//*[@id="mainView"]/div/div[2]/div/div/button')
	#button.click()
	time.sleep(5)
	driver.quit()
	return

if __name__ == '__main__':
    main()
